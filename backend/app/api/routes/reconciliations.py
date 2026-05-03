import io
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from typing import List, Any, Dict, Tuple, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
from fastapi.concurrency import run_in_threadpool
from datetime import datetime, timezone
from app.core.database import get_db
from app.models.user import User
from app.models.organization import Membership
from app.schemas.reconciliation import AnomalyResolveRequest
from app.models.reconciliation import (
    BankAccount,
    Reconciliation,
    ReconciliationStatus,
    BankTransaction,
    LedgerTransaction,
    Anomaly,
)
from app.models.match import Match
from app.schemas.reconciliation import (
    BankAccountCreate,
    BankAccountResponse,
    ReconciliationCreate,
    ReconciliationResponse,
    FileUploadResponse,
)
from app.schemas.transaction import (
    BankTransactionRead,
    LedgerTransactionRead,
)
from app.schemas.match import MatchRead
from app.api.deps import get_current_user
from app.services.file_parser import parse_bank_file, parse_ledger_file
from app.services.matching_engine import run_matching
from app.services.anomaly_agent import run_anomaly_detection  # Imported Anomaly Agent
from app.schemas.reconciliation import AnomalyRead
router = APIRouter()


@router.get("/", response_model=List[ReconciliationResponse])
async def list_reconciliations(
    org_id: Optional[int] = Query(None),
    bank_account_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List reconciliations, optionally filtered by org, bank account, or status."""
    stmt = (
        select(Reconciliation)
        .join(BankAccount, BankAccount.id == Reconciliation.bank_account_id)
        .where(Reconciliation.created_by_id == current_user.id)
    )
    if org_id is not None:
        stmt = stmt.where(BankAccount.organization_id == org_id)
    if bank_account_id is not None:
        stmt = stmt.where(Reconciliation.bank_account_id == bank_account_id)
    if status is not None:
        stmt = stmt.where(Reconciliation.status == status)
    stmt = stmt.order_by(Reconciliation.created_at.desc())
    result = await db.execute(stmt)
    return result.scalars().all()


async def _get_reconciliation_or_404(
    db: AsyncSession,
    reconciliation_id: int,
    user_id: int,
) -> Reconciliation:
    stmt = (
        select(Reconciliation)
        .join(BankAccount, BankAccount.id == Reconciliation.bank_account_id)
        .where(
            Reconciliation.id == reconciliation_id,
            Reconciliation.created_by_id == user_id,
        )
    )
    result = await db.execute(stmt)
    rec = result.scalar_one_or_none()
    if not rec:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Reconciliation not found",
        )
    return rec


async def _get_transactions_for_reconciliation(
    db: AsyncSession,
    reconciliation_id: int,
) -> Tuple[List[BankTransaction], List[LedgerTransaction]]:
    bank_stmt = (
        select(BankTransaction)
        .where(BankTransaction.reconciliation_id == reconciliation_id)
        .order_by(BankTransaction.transaction_date.asc())
    )
    ledger_stmt = (
        select(LedgerTransaction)
        .where(LedgerTransaction.reconciliation_id == reconciliation_id)
        .order_by(LedgerTransaction.transaction_date.asc())
    )

    bank_result = await db.execute(bank_stmt)
    ledger_result = await db.execute(ledger_stmt)

    bank_txs = bank_result.scalars().all()
    ledger_txs = ledger_result.scalars().all()
    return bank_txs, ledger_txs


@router.post(
    "/bank-accounts",
    response_model=BankAccountResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_bank_account(
    org_id: int,
    account_data: BankAccountCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Create a new bank account for reconciliation."""
    result = await db.execute(
        select(Membership)
        .where(Membership.user_id == current_user.id)
        .where(Membership.organization_id == org_id)
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=403, detail="Not a member of this organization")

    new_account = BankAccount(
        organization_id=org_id,
        **account_data.model_dump(),
    )
    db.add(new_account)
    await db.commit()
    await db.refresh(new_account)
    return new_account


@router.get("/bank-accounts", response_model=List[BankAccountResponse])
async def list_bank_accounts(
    org_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List all bank accounts for an organization."""
    result = await db.execute(
        select(BankAccount).where(BankAccount.organization_id == org_id)
    )
    return result.scalars().all()


@router.post("/upload/bank", response_model=FileUploadResponse)
async def upload_bank_file(
    reconciliation_id: int = Form(...),
    file: UploadFile = File(...),
    date_column: str = Form("Date"),
    amount_column: str = Form("Amount"),
    description_column: str = Form("Description"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Upload and parse a bank statement file (CSV/Excel)."""
    result = await db.execute(
        select(Reconciliation).where(Reconciliation.id == reconciliation_id)
    )
    reconciliation = result.scalar_one_or_none()
    if not reconciliation:
        raise HTTPException(status_code=404, detail="Reconciliation not found")

    # --- MEMORY PROTECTION: File Size Check ---
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB limit
    file.file.seek(0, 2)
    if file.file.tell() > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large. Max size is 10MB.")
    await file.seek(0)
    # ------------------------------------------

    try:
        transactions = await parse_bank_file(
            file=file,
            date_column=date_column,
            amount_column=amount_column,
            description_column=description_column,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File parsing error: {str(e)}")

    for tx_data in transactions:
        tx = BankTransaction(
            reconciliation_id=reconciliation_id,
            transaction_date=tx_data["date"],
            amount=tx_data["amount"],
            description=tx_data.get("description"),
            raw_data=str(tx_data),
        )
        db.add(tx)

    reconciliation.total_bank_transactions = len(transactions)
    await db.commit()

    return FileUploadResponse(
        message="Bank file uploaded successfully",
        rows_parsed=len(transactions),
        file_type="bank",
    )


@router.post("/upload/ledger", response_model=FileUploadResponse)
async def upload_ledger_file(
    reconciliation_id: int = Form(...),
    file: UploadFile = File(...),
    date_column: str = Form("Date"),
    debit_column: str = Form("Debit"),
    credit_column: str = Form("Credit"),
    description_column: str = Form("Description"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Upload and parse a ledger file (CSV/Excel)."""
    result = await db.execute(
        select(Reconciliation).where(Reconciliation.id == reconciliation_id)
    )
    reconciliation = result.scalar_one_or_none()
    if not reconciliation:
        raise HTTPException(status_code=404, detail="Reconciliation not found")

    # --- MEMORY PROTECTION: File Size Check ---
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB limit
    file.file.seek(0, 2)
    if file.file.tell() > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large. Max size is 10MB.")
    await file.seek(0)
    # ------------------------------------------

    try:
        transactions = await parse_ledger_file(
            file=file,
            date_column=date_column,
            debit_column=debit_column,
            credit_column=credit_column,
            description_column=description_column,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File parsing error: {str(e)}")

    for tx_data in transactions:
        tx = LedgerTransaction(
            reconciliation_id=reconciliation_id,
            transaction_date=tx_data["date"],
            debit=tx_data.get("debit", 0),
            credit=tx_data.get("credit", 0),
            description=tx_data.get("description"),
            raw_data=str(tx_data),
        )
        db.add(tx)

    reconciliation.total_ledger_transactions = len(transactions)
    await db.commit()

    return FileUploadResponse(
        message="Ledger file uploaded successfully",
        rows_parsed=len(transactions),
        file_type="ledger",
    )


@router.post(
    "/",
    response_model=ReconciliationResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_reconciliation(
    recon_data: ReconciliationCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Create a new reconciliation session."""
    new_recon = Reconciliation(
        bank_account_id=recon_data.bank_account_id,
        created_by_id=current_user.id,
        period_start=recon_data.period_start,
        period_end=recon_data.period_end,
    )
    db.add(new_recon)
    await db.commit()
    await db.refresh(new_recon)
    return new_recon


@router.get("/{recon_id}", response_model=ReconciliationResponse)
async def get_reconciliation_basic(
    recon_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get reconciliation details."""
    result = await db.execute(
        select(Reconciliation).where(Reconciliation.id == recon_id)
    )
    recon = result.scalar_one_or_none()
    if not recon:
        raise HTTPException(status_code=404, detail="Reconciliation not found")
    return recon


@router.get(
    "/{reconciliation_id}/transactions",
    response_model=Dict[str, Any],
)
async def get_reconciliation_transactions(
    reconciliation_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Dict[str, Any]:
    rec = await _get_reconciliation_or_404(db, reconciliation_id, current_user.id)

    bank_txs, ledger_txs = await _get_transactions_for_reconciliation(db, rec.id)

    return {
        "reconciliation_id": rec.id,
        "bank_transactions": [BankTransactionRead.model_validate(tx) for tx in bank_txs],
        "ledger_transactions": [LedgerTransactionRead.model_validate(tx) for tx in ledger_txs],
    }


@router.post(
    "/{reconciliation_id}/match",
    response_model=Dict[str, Any],
)
async def run_reconciliation_matching(
    reconciliation_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Dict[str, Any]:
    rec = await _get_reconciliation_or_404(db, reconciliation_id, current_user.id)

    bank_txs, ledger_txs = await _get_transactions_for_reconciliation(db, rec.id)

    if not bank_txs or not ledger_txs:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Both bank and ledger transactions are required to run matching.",
        )

    rec.total_bank_transactions = len(bank_txs)
    rec.total_ledger_transactions = len(ledger_txs)

    # 1. Clear old matches and anomalies for this reconciliation
    await db.execute(delete(Match).where(Match.reconciliation_id == rec.id))
    await db.execute(delete(Anomaly).where(Anomaly.reconciliation_id == rec.id))

    # 2. --- RUN MATCHING ENGINE (Background Thread) ---
    match_dicts, unmatched_bank_ids, unmatched_ledger_ids = await run_in_threadpool(
        run_matching,
        reconciliation_id=rec.id,
        bank_transactions=bank_txs,
        ledger_transactions=ledger_txs,
    )
    
    match_models = [Match(**m) for m in match_dicts]
    db.add_all(match_models)

    rec.matched_count = len(match_models)
    rec.unmatched_bank_count = len(unmatched_bank_ids)
    rec.unmatched_ledger_count = len(unmatched_ledger_ids)
    
    await db.commit() # Save matches first so Anomaly agent can see them if needed

    # 3. --- RUN ANOMALY AGENT (Background Thread) ---
    anomaly_dicts = await run_in_threadpool(
        run_anomaly_detection,
        reconciliation_id=rec.id,
        bank_transactions=bank_txs,
        ledger_transactions=ledger_txs,
        matches=match_dicts
    )

    anomaly_models = [Anomaly(**a) for a in anomaly_dicts]
    if anomaly_models:
        db.add_all(anomaly_models)

    rec.anomaly_count = len(anomaly_models)
    rec.status = ReconciliationStatus.COMPLETED
    rec.completed_at = datetime.now(timezone.utc)

    await db.commit()
    await db.refresh(rec)

    return {
        "reconciliation_id": rec.id,
        "matched_count": rec.matched_count,
        "unmatched_bank_count": rec.unmatched_bank_count,
        "unmatched_ledger_count": rec.unmatched_ledger_count,
        "anomaly_count": rec.anomaly_count,
        "matches": [MatchRead.model_validate(m) for m in match_models],
        "unmatched_bank_ids": unmatched_bank_ids,
        "unmatched_ledger_ids": unmatched_ledger_ids,
    }


@router.get(
    "/{reconciliation_id}/matches",
    response_model=List[MatchRead],
)
async def get_reconciliation_matches(
    reconciliation_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> List[Match]:
    rec = await _get_reconciliation_or_404(db, reconciliation_id, current_user.id)

    stmt = (
        select(Match)
        .where(Match.reconciliation_id == rec.id)
        .order_by(Match.created_at.asc())
    )
    result = await db.execute(stmt)
    matches = result.scalars().all()
    return matches

@router.get(
    "/{reconciliation_id}/anomalies",
    response_model=List[AnomalyRead],
)
async def get_reconciliation_anomalies(
    reconciliation_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Fetch all detected anomalies for a reconciliation session."""
    rec = await _get_reconciliation_or_404(db, reconciliation_id, current_user.id)

    stmt = (
        select(Anomaly)
        .where(Anomaly.reconciliation_id == rec.id)
        .order_by(Anomaly.severity.desc()) # Show HIGH severity first
    )
    result = await db.execute(stmt)
    return result.scalars().all()


@router.post("/resolve-anomaly", response_model=AnomalyRead)
async def resolve_anomaly(
    resolution_data: AnomalyResolveRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    The Automation Engine: Executes the Copilot's suggested actions to fix the database.
    """
    # 1. Fetch the Anomaly to ensure it exists and isn't already resolved
    stmt = select(Anomaly).where(Anomaly.id == resolution_data.anomaly_id)
    result = await db.execute(stmt)
    anomaly = result.scalar_one_or_none()

    if not anomaly:
        raise HTTPException(status_code=404, detail="Anomaly not found")
    if anomaly.is_resolved:
        raise HTTPException(status_code=400, detail="Anomaly is already resolved")

    resolution_note = resolution_data.note or f"Resolved via Copilot Action: {resolution_data.action}"

    if resolution_data.action == "create_journal_entry":
        # Missing in ledger? Let's create the compensating ledger transaction
        if not resolution_data.amount:
             raise HTTPException(status_code=400, detail="Amount is required for create_journal_entry")
             
        new_ledger_tx = LedgerTransaction(
            reconciliation_id=anomaly.reconciliation_id,
            transaction_date=datetime.now(timezone.utc).date(), # Record it today
            debit=abs(resolution_data.amount) if resolution_data.amount > 0 else 0,
            credit=abs(resolution_data.amount) if resolution_data.amount < 0 else 0,
            description=f"AI Copilot Auto-Adjustment for Anomaly #{anomaly.id}",
            raw_data='{"source": "system_generated", "type": "copilot_fix"}'
        )
        db.add(new_ledger_tx)
        resolution_note += f" | Created Ledger Entry for {resolution_data.amount}"

    elif resolution_data.action == "delete_duplicate":
        # Duplicate in ledger? Let's delete the exact offending transaction
        if anomaly.ledger_transaction_id:
            tx_id = anomaly.ledger_transaction_id
            
            # 1. Disconnect the Foreign Key in memory
            anomaly.ledger_transaction_id = None 
            
            # 2. THE FIX: Flush the update to Postgres immediately!
            await db.flush()
            
            await db.execute(delete(LedgerTransaction).where(LedgerTransaction.id == tx_id))
            resolution_note += f" | Deleted duplicate ledger transaction #{tx_id}."
        else:
            raise HTTPException(status_code=400, detail="No ledger transaction ID attached to this anomaly.")
    elif resolution_data.action == "reverse_transaction":

        resolution_note += " | Flagged for manual reversal by accounting team."
    
    else:
        # Fallback for manual review
        resolution_note += " | Marked as reviewed."


    anomaly.is_resolved = True
    anomaly.resolved_by_id = current_user.id
    anomaly.resolved_at = datetime.now(timezone.utc)
    anomaly.resolution_note = resolution_note


    await db.commit()
    await db.refresh(anomaly)

    return anomaly

@router.get("/{reconciliation_id}/export")
async def export_audit_report(
    reconciliation_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Generate a color-coded 'Turnitin-style' Excel audit report."""
    rec = await _get_reconciliation_or_404(db, reconciliation_id, current_user.id)
    bank_txs, ledger_txs = await _get_transactions_for_reconciliation(db, rec.id)


    match_result = await db.execute(select(Match).where(Match.reconciliation_id == rec.id))
    matches = match_result.scalars().all()
    
    anomaly_result = await db.execute(select(Anomaly).where(Anomaly.reconciliation_id == rec.id))
    anomalies = anomaly_result.scalars().all()


    matched_bank_ids = {m.bank_transaction_id for m in matches if m.bank_transaction_id}
    matched_ledger_ids = {m.ledger_transaction_id for m in matches if m.ledger_transaction_id}
    
    anomaly_by_bank = {a.bank_transaction_id: a for a in anomalies if a.bank_transaction_id}
    anomaly_by_ledger = {a.ledger_transaction_id: a for a in anomalies if a.ledger_transaction_id}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Audit_Report_{rec.id}"


    GREEN_FILL = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid") # Matched perfectly
    RED_FILL = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")     # Unresolved Anomaly
    YELLOW_FILL = PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid")  # Resolved via Copilot


    headers = ["Type", "Status", "Date", "Description", "Amount", "Copilot Notes"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)


    for tx in bank_txs:
        status = "Matched"
        fill = GREEN_FILL
        notes = "Exact match found."
        comment_text = None

        if tx.id in anomaly_by_bank:
            anomaly = anomaly_by_bank[tx.id]
            if anomaly.is_resolved:
                status = f"Resolved: {anomaly.anomaly_type}"
                fill = YELLOW_FILL
                notes = anomaly.resolution_note
                comment_text = f"Copilot Analysis: {anomaly.description}\nResolution: {anomaly.resolution_note}"
            else:
                status = f"ISSUE: {anomaly.anomaly_type}"
                fill = RED_FILL
                notes = anomaly.suggested_action
                comment_text = f"AI WARNING: {anomaly.title}\n{anomaly.description}"
        elif tx.id not in matched_bank_ids:
            status = "Unmatched"
            fill = RED_FILL
            notes = "No matching ledger entry found by AI or rules."

        row_data = ["Bank", status, str(tx.transaction_date), tx.description, tx.amount, notes]
        ws.append(row_data)
        
        current_row = ws.max_row
        for col_idx in range(1, len(row_data) + 1):
            ws.cell(row=current_row, column=col_idx).fill = fill
            

        if comment_text:
            comment = Comment(comment_text, "ReFinely AI Copilot")
            comment.width = 300
            comment.height = 100
            ws.cell(row=current_row, column=2).comment = comment # Attach comment to Status column


    for tx in ledger_txs:
        # If it's a perfect match, we don't need to list it twice. Just list the problems.
        if tx.id in matched_ledger_ids and tx.id not in anomaly_by_ledger:
            continue 

        status = "Unmatched"
        fill = RED_FILL
        notes = "No matching bank entry found."
        comment_text = None

        if tx.id in anomaly_by_ledger:
            anomaly = anomaly_by_ledger[tx.id]
            if anomaly.is_resolved:
                status = f"Resolved: {anomaly.anomaly_type}"
                fill = YELLOW_FILL
                notes = anomaly.resolution_note
            else:
                status = f"ISSUE: {anomaly.anomaly_type}"
                fill = RED_FILL
                notes = anomaly.suggested_action
                comment_text = f"AI WARNING: {anomaly.title}\n{anomaly.description}"

        amount = tx.debit if tx.debit > 0 else -tx.credit
        row_data = ["Ledger", status, str(tx.transaction_date), tx.description, amount, notes]
        ws.append(row_data)

        current_row = ws.max_row
        for col_idx in range(1, len(row_data) + 1):
            ws.cell(row=current_row, column=col_idx).fill = fill
            
        if comment_text:
            comment = Comment(comment_text, "ReFinely AI Copilot")
            comment.width = 300
            comment.height = 100
            ws.cell(row=current_row, column=2).comment = comment

    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['F'].width = 60

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=ReFinely_Audit_{rec.id}.xlsx"
        }
    )